"""
AI Email Sender v4 — Draft → Edit → Send Flow
==============================================
The correct workflow:

  Step 1:  python send_emails.py --draft
           AI generates all emails and saves them to drafts.json
           You edit drafts.json to tweak any email you want

  Step 2:  python send_emails.py --test
           Reads YOUR edited drafts.json
           Sends everything to TEST_EMAIL so you can verify in your inbox

  Step 3:  python send_emails.py --send
           Reads the SAME drafts.json (your edited version)
           Sends to real recipients with delays

  Optional: python send_emails.py --check-replies
           Scans Gmail inbox for replies, updates Excel status

Key principle: The AI runs ONCE (--draft). Every subsequent mode
reads from YOUR edited file — never re-generates.
"""

import os
import csv
import sys
import time
import random
import smtplib
import imaplib
import logging
import json
import re
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import ollama
from openpyxl import load_workbook
from dotenv import load_dotenv

# ── Config ─────────────────────────────────────────────────────────────────────

load_dotenv()

EXCEL_FILE      = "contacts.xlsx"
CONTEXT_FILE    = "context.md"
ATTACHMENTS_DIR = "attachments"
LOG_FILE        = "sent_log.csv"
DRAFTS_FILE     = "drafts.json"          # machine-readable, you edit this
PREVIEW_FILE    = "drafts_preview.txt"   # human-readable preview (auto-generated from drafts.json)

MODEL          = os.getenv("OLLAMA_MODEL", "gemma4:e4b")
GMAIL_USER     = os.getenv("GMAIL_USER")
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASSWORD")
SENDER_NAME    = os.getenv("SENDER_NAME", "Your Name")
TEST_EMAIL     = os.getenv("TEST_EMAIL", "")

DELAY_MIN    = int(os.getenv("DELAY_MIN_SECONDS", "30"))
DELAY_MAX    = int(os.getenv("DELAY_MAX_SECONDS", "60"))
DAILY_CAP    = int(os.getenv("DAILY_SEND_CAP", "200"))
MAX_RETRIES  = int(os.getenv("MAX_RETRIES", "3"))
DRAFT_WORKERS = int(os.getenv("DRAFT_WORKERS", "3"))

STATUS_SENT_ONCE  = "email sent"
STATUS_SENT_TWICE = "sent twice"
STATUS_FAILED     = "failed"
STATUS_REPLIED    = "replied"

NO_ATTACHMENT_VALUES = {"", "na", "n/a", "none", "no", "nan", "-"}
EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)

# ── Helpers ────────────────────────────────────────────────────────────────────

def load_context():
    p = Path(CONTEXT_FILE)
    if not p.exists():
        log.error(f"'{CONTEXT_FILE}' not found. Create it first (see README).")
        return None
    return p.read_text(encoding="utf-8").strip()


def list_attachments():
    p = Path(ATTACHMENTS_DIR)
    return [f.name for f in p.iterdir() if f.is_file()] if p.exists() else []


def is_valid_email(email):
    return bool(EMAIL_REGEX.match(email.strip()))


def needs_no_attachment(hint):
    return hint.lower().strip() in NO_ATTACHMENT_VALUES


def next_status(current):
    s = (current or "").lower().strip()
    return STATUS_SENT_TWICE if s == STATUS_SENT_ONCE else STATUS_SENT_ONCE


def should_skip(current_status):
    s = (current_status or "").lower().strip()
    if s == STATUS_SENT_TWICE:
        return True, "Already sent twice"
    if s == STATUS_REPLIED:
        return True, "Already replied"
    return False, ""


# ── AI calls ───────────────────────────────────────────────────────────────────

def match_file(hint, filenames):
    if not hint or not filenames:
        return None
    file_list = "\n".join(f"- {f}" for f in filenames)
    prompt = f"""You are a file-matching assistant.

Available files:
{file_list}

The contact needs: "{hint}"

Reply with ONLY the exact filename, or 'null' if nothing matches."""

    try:
        r = ollama.chat(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            options={"temperature": 0}
        )
        result = r["message"]["content"].strip().strip('"').strip("'")
        return result if result in filenames else None
    except Exception as e:
        log.error(f"File match error: {e}")
        return None


def draft_email_ai(context, name, description, attached_file, attempt_type):
    attachment_note = (
        f"\nAttachment: '{attached_file}' — reference it naturally."
        if attached_file else
        "\nAttachment: None — do not mention any attachment."
    )
    followup_note = ""
    if attempt_type == "followup":
        followup_note = """
=== THIS IS A FOLLOW-UP ===
Write a short, polite follow-up (2-3 sentences).
Acknowledge you reached out before. Don't be pushy.
Do NOT copy the first email."""

    prompt = f"""You are writing an email on behalf of {SENDER_NAME}.

=== CAMPAIGN CONTEXT ===
{context}

=== RECIPIENT ===
Name: {name}
Details: {description}
{attachment_note}
{followup_note}

Sign off with: {SENDER_NAME}

Reply with JSON only — no markdown, no extra text:
{{"subject": "subject line here", "body": "email body here with \\n for line breaks"}}"""

    r = ollama.chat(
        model=MODEL,
        messages=[{"role": "user", "content": prompt}],
        options={"temperature": 0.5}
    )
    text = re.sub(r"```json|```", "", r["message"]["content"]).strip()
    try:
        parsed = json.loads(text)
        return parsed.get("subject", "Following up"), parsed.get("body", text)
    except json.JSONDecodeError:
        return "Following up", text


# ── Drafts file I/O ────────────────────────────────────────────────────────────
#
# drafts.json structure:
# [
#   {
#     "email":          "priya@stripe.com",
#     "name":           "Priya Sharma",
#     "attempt_type":   "first",
#     "matched_file":   "resume_2024.pdf",
#     "attachment_path": "attachments/resume_2024.pdf",
#     "subject":        "Your subject here",    ← YOU EDIT THESE
#     "body":           "Your body here",       ← YOU EDIT THESE
#     "row_idx":        2,
#     "file_hint":      "resume"
#   },
#   ...
# ]
#
# To edit: open drafts.json in any text editor, change "subject" and "body"
# fields. Save. Then run --test or --send.

def save_drafts(drafts):
    """Save drafts to JSON — machine-readable, human-editable."""
    # Strip attachment_path from what we save (it gets re-resolved on send)
    clean = []
    for d in drafts:
        clean.append({
            "email":           d["email"],
            "name":            d["name"],
            "attempt_type":    d["attempt_type"],
            "matched_file":    d.get("matched_file") or "",
            "attachment_path": d.get("attachment_path") or "",
            "subject":         d["subject"],
            "body":            d["body"],
            "row_idx":         d["_row_idx"],
            "file_hint":       d.get("file_hint", ""),
            "current_status":  d.get("status", ""),
        })
    with open(DRAFTS_FILE, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=2, ensure_ascii=False)
    log.info(f"✓ Saved {len(clean)} drafts to '{DRAFTS_FILE}'")


def load_drafts():
    """Load drafts from JSON — includes any edits you made."""
    if not Path(DRAFTS_FILE).exists():
        log.error(f"'{DRAFTS_FILE}' not found. Run --draft first.")
        return None
    with open(DRAFTS_FILE, "r", encoding="utf-8") as f:
        drafts = json.load(f)
    log.info(f"✓ Loaded {len(drafts)} drafts from '{DRAFTS_FILE}'")
    return drafts


def write_preview(drafts):
    """Write a human-readable .txt preview from the current drafts.json state."""
    with open(PREVIEW_FILE, "w", encoding="utf-8") as f:
        f.write(f"EMAIL PREVIEW — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("Edit drafts.json to change any email. This file is regenerated automatically.\n")
        f.write("=" * 70 + "\n\n")
        for d in drafts:
            f.write(f"To:      {d['name']} <{d['email']}>\n")
            f.write(f"Type:    {d['attempt_type']}\n")
            f.write(f"Attach:  {d.get('matched_file') or '(none)'}\n")
            f.write(f"Subject: {d['subject']}\n")
            f.write("-" * 70 + "\n")
            body = d["body"].replace("\\n", "\n")
            f.write(body + "\n")
            f.write("=" * 70 + "\n\n")
    log.info(f"✓ Human-readable preview saved to '{PREVIEW_FILE}'")


# ── Excel I/O ─────────────────────────────────────────────────────────────────

def read_contacts():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [str(c.value).lower().strip() if c.value else "" for c in ws[1]]

    contacts = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        data = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                data[headers[i]] = str(cell.value).strip() if cell.value is not None else ""
        data["_row_idx"] = row_idx
        contacts.append(data)
    return wb, ws, headers, contacts


def ensure_status_column(wb, ws, headers):
    if "status" not in headers:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value="status")
        headers.append("status")
        wb.save(EXCEL_FILE)
        log.info("Added 'status' column to spreadsheet")
    return headers.index("status") + 1


def update_status(wb, ws, row_idx, status_col, value):
    ws.cell(row=row_idx, column=status_col, value=value)
    wb.save(EXCEL_FILE)


# ── CSV log ───────────────────────────────────────────────────────────────────

def init_log():
    if not Path(LOG_FILE).exists():
        with open(LOG_FILE, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                "timestamp", "name", "email", "attempt_type",
                "matched_file", "subject", "status", "note"
            ])


def write_log(entry):
    with open(LOG_FILE, "a", newline="", encoding="utf-8") as f:
        csv.DictWriter(f, fieldnames=[
            "timestamp", "name", "email", "attempt_type",
            "matched_file", "subject", "status", "note"
        ]).writerow(entry)


# ── SMTP send ─────────────────────────────────────────────────────────────────

def send_email_with_retry(to_email, subject, body, attachment_path):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            msg = MIMEMultipart()
            msg["From"]    = f"{SENDER_NAME} <{GMAIL_USER}>"
            msg["To"]      = to_email
            msg["Subject"] = subject
            msg.attach(MIMEText(body.replace("\\n", "\n"), "plain", "utf-8"))

            if attachment_path and Path(attachment_path).exists():
                path = Path(attachment_path)
                with open(path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{path.name}"')
                msg.attach(part)

            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(GMAIL_USER, GMAIL_APP_PASS)
                server.sendmail(GMAIL_USER, to_email, msg.as_string())
            return True, None

        except smtplib.SMTPException as e:
            err = str(e)
            if any(code in err for code in ["421", "454", "550"]):
                return False, f"Gmail throttling: {err}"
            backoff = 2 ** attempt
            log.warning(f"   ↻ Retry {attempt}/{MAX_RETRIES} in {backoff}s — {err}")
            time.sleep(backoff)
        except Exception as e:
            backoff = 2 ** attempt
            log.warning(f"   ↻ Retry {attempt}/{MAX_RETRIES} in {backoff}s — {e}")
            time.sleep(backoff)
    return False, "Max retries exceeded"


# ── IMAP reply check ──────────────────────────────────────────────────────────

def check_replies(contact_emails):
    log.info("Checking Gmail for replies…")
    replied = set()
    try:
        with imaplib.IMAP4_SSL("imap.gmail.com") as mail:
            mail.login(GMAIL_USER, GMAIL_APP_PASS)
            mail.select('"[Gmail]/All Mail"')
            date_since = (datetime.now() - timedelta(days=30)).strftime("%d-%b-%Y")
            for addr in contact_emails:
                status, data = mail.search(None, f'(FROM "{addr}" SINCE {date_since})')
                if status == "OK" and data[0]:
                    replied.add(addr.lower())
    except Exception as e:
        log.warning(f"IMAP check failed: {e}")
        return set()
    log.info(f"Found {len(replied)} replies")
    return replied


# ── Countdown timer ───────────────────────────────────────────────────────────

def countdown_delay(seconds):
    for r in range(seconds, 0, -1):
        print(f"\r   ⏳ Next email in {r}s... ", end="", flush=True)
        time.sleep(1)
    print("\r" + " " * 50 + "\r", end="", flush=True)


# ── Mode handlers ─────────────────────────────────────────────────────────────

def run_draft():
    """
    --draft mode
    Runs AI on every eligible contact, saves results to drafts.json
    and a human-readable drafts_preview.txt.
    You edit drafts.json before running --test or --send.
    """
    log.info("MODE: DRAFT — AI will generate emails and save to drafts.json")
    log.info("After this, edit drafts.json then run --test or --send")
    log.info("=" * 60)

    context = load_context()
    if not context:
        return

    try:
        wb, ws, headers, all_contacts = read_contacts()
    except FileNotFoundError:
        log.error(f"'{EXCEL_FILE}' not found")
        return

    status_col = ensure_status_column(wb, ws, headers)
    available_files = list_attachments()

    # Filter eligible contacts
    seen = set()
    to_draft = []
    for c in all_contacts:
        email = c.get("email", "").strip().lower()
        if not email or email == "nan":
            continue
        if not is_valid_email(email):
            log.warning(f"⚠ Invalid email skipped: {c.get('email')}")
            continue
        if email in seen:
            log.warning(f"⚠ Duplicate skipped: {email}")
            continue
        seen.add(email)
        skip, reason = should_skip(c.get("status", ""))
        if skip:
            log.info(f"   ⊘ {email} — {reason}")
            continue
        to_draft.append(c)

    if not to_draft:
        log.info("No contacts need drafting.")
        return

    log.info(f"Drafting {len(to_draft)} emails using {MODEL}…")

    def make_draft(c):
        try:
            status       = c.get("status", "")
            attempt_type = "followup" if status.lower() == STATUS_SENT_ONCE else "first"
            file_hint    = c.get("file_hint", "") or c.get("file", "")
            matched_file    = None
            attachment_path = None

            if not needs_no_attachment(file_hint) and available_files:
                matched_file = match_file(file_hint, available_files)
                if matched_file:
                    attachment_path = str(Path(ATTACHMENTS_DIR) / matched_file)

            subject, body = draft_email_ai(
                context, c.get("name", "there"),
                c.get("description", ""), matched_file, attempt_type
            )
            return {**c, "attempt_type": attempt_type, "matched_file": matched_file,
                    "attachment_path": attachment_path, "subject": subject,
                    "body": body, "draft_ok": True}
        except Exception as e:
            return {**c, "draft_ok": False, "draft_error": str(e)}

    drafts = []
    with ThreadPoolExecutor(max_workers=DRAFT_WORKERS) as executor:
        futures = {executor.submit(make_draft, c): c for c in to_draft}
        for i, fut in enumerate(as_completed(futures), start=1):
            d = fut.result()
            if d.get("draft_ok"):
                tag = "📎" if d.get("matched_file") else "📭"
                log.info(f"   [{i}/{len(to_draft)}] {tag} {d['name']} — \"{d['subject']}\"")
            else:
                log.error(f"   [{i}/{len(to_draft)}] ✗ {d.get('name')} — {d.get('draft_error')}")
            drafts.append(d)

    # Restore original sheet order
    order = {c["email"].lower(): idx for idx, c in enumerate(to_draft)}
    drafts.sort(key=lambda d: order.get(d.get("email", "").lower(), 9999))

    good = [d for d in drafts if d.get("draft_ok")]

    if not good:
        log.error("All drafts failed. Check Ollama is running.")
        return

    save_drafts(good)
    write_preview(good)

    log.info("=" * 60)
    log.info(f"Done. {len(good)} drafts saved.")
    log.info(f"")
    log.info(f"  → Open 'drafts.json' and edit any subject or body you want")
    log.info(f"  → Open 'drafts_preview.txt' for easy reading")
    log.info(f"  → Then run:  python send_emails.py --test")
    log.info(f"  → Then run:  python send_emails.py --send")
    log.info("=" * 60)


def run_send(test_mode=False):
    """
    --test and --send mode
    Both read from drafts.json (your edited version).
    --test  sends to TEST_EMAIL, does not update Excel status
    --send  sends to real recipients, updates Excel status
    """
    mode_label = "TEST (sending to yourself)" if test_mode else "SEND (sending to real recipients)"
    log.info(f"MODE: {mode_label}")
    log.info("Reading from drafts.json — your edits are respected")
    log.info("=" * 60)

    if not GMAIL_USER or not GMAIL_APP_PASS:
        log.error("Missing GMAIL_USER or GMAIL_APP_PASSWORD in .env")
        return

    if test_mode and not TEST_EMAIL:
        log.error("--test mode needs TEST_EMAIL set in .env")
        return

    drafts = load_drafts()
    if not drafts:
        return

    # For --send, load Excel to update status
    wb = ws = status_col = None
    if not test_mode:
        try:
            wb, ws, headers, _ = read_contacts()
            status_col = ensure_status_column(wb, ws, headers)
        except FileNotFoundError:
            log.error(f"'{EXCEL_FILE}' not found")
            return

    # For --send, check replies first so we skip people who already wrote back
    if not test_mode:
        try:
            replied = check_replies([d["email"] for d in drafts])
            if replied:
                remaining = []
                for d in drafts:
                    if d["email"].lower() in replied:
                        log.info(f"   ↩ {d['name']} already replied — skipping")
                        update_status(wb, ws, d["row_idx"], status_col, STATUS_REPLIED)
                    else:
                        remaining.append(d)
                drafts = remaining
        except Exception as e:
            log.warning(f"Reply check skipped: {e}")

    if not drafts:
        log.info("No emails to send after reply check.")
        return

    init_log()
    stats  = {"sent": 0, "failed": 0}
    sent_today = 0

    log.info(f"Sending {len(drafts)} emails…")
    if not test_mode:
        log.info(f"Delays: {DELAY_MIN}-{DELAY_MAX}s | Daily cap: {DAILY_CAP}")
    log.info("=" * 60)

    for idx, d in enumerate(drafts):
        if not test_mode and sent_today >= DAILY_CAP:
            log.warning(f"Daily cap of {DAILY_CAP} reached. Run again tomorrow.")
            break

        target = TEST_EMAIL if test_mode else d["email"]
        log.info(f"[{idx+1}/{len(drafts)}] → {d['name']} <{target}>")
        if test_mode:
            log.info(f"   (real recipient: {d['email']})")

        attachment_path = d.get("attachment_path") or None

        success, err = send_email_with_retry(
            target, d["subject"], d["body"], attachment_path
        )

        log_entry = {
            "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "name":         d["name"],
            "email":        d["email"],
            "attempt_type": d["attempt_type"],
            "matched_file": d.get("matched_file") or "",
            "subject":      d["subject"],
            "status":       "",
            "note":         "",
        }

        if success:
            log.info(f"   ✓ Sent — \"{d['subject']}\"")
            log_entry["status"] = "SENT"
            stats["sent"] += 1
            sent_today += 1
            if not test_mode:
                new_status = next_status(d.get("current_status", ""))
                update_status(wb, ws, d["row_idx"], status_col, new_status)
        else:
            log.error(f"   ✗ Failed: {err}")
            log_entry["status"] = "FAILED"
            log_entry["note"]   = err
            stats["failed"] += 1
            if not test_mode:
                update_status(wb, ws, d["row_idx"], status_col, STATUS_FAILED)
            if err and any(x in err for x in ["throttling", "421", "454"]):
                log.error("Gmail throttling detected — stopping.")
                write_log(log_entry)
                break

        write_log(log_entry)

        if idx < len(drafts) - 1 and (not test_mode) and sent_today < DAILY_CAP:
            countdown_delay(random.randint(DELAY_MIN, DELAY_MAX))

    log.info("=" * 60)
    log.info(f"Done. Sent: {stats['sent']}  Failed: {stats['failed']}")
    if not test_mode:
        log.info("Excel status column updated.")
    log.info(f"Log: '{LOG_FILE}'")
    log.info("=" * 60)


def run_check_replies():
    """--check-replies mode: scan inbox and update Excel."""
    log.info("MODE: CHECK REPLIES")
    log.info("=" * 60)

    try:
        wb, ws, headers, contacts = read_contacts()
        status_col = ensure_status_column(wb, ws, headers)
    except FileNotFoundError:
        log.error(f"'{EXCEL_FILE}' not found")
        return

    emails = [c.get("email", "") for c in contacts if c.get("email")]
    replied = check_replies(emails)

    for c in contacts:
        if c.get("email", "").lower() in replied:
            update_status(wb, ws, c["_row_idx"], status_col, STATUS_REPLIED)
            log.info(f"   ↩ Marked as replied: {c.get('name')} <{c.get('email')}>")

    log.info("=" * 60)
    log.info("Done. Excel updated.")
    log.info("=" * 60)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="AI Email Sender — Draft → Edit → Send",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Workflow:
  1. python send_emails.py --draft          # AI generates, saves drafts.json
  2. Edit drafts.json (change subject/body)
  3. python send_emails.py --test           # sends your edits to TEST_EMAIL
  4. python send_emails.py --send           # sends your edits to real recipients
  5. python send_emails.py --check-replies  # scan inbox, update Excel
        """
    )
    parser.add_argument("--draft",         action="store_true", help="Generate drafts with AI and save to drafts.json")
    parser.add_argument("--test",          action="store_true", help="Send drafts.json to TEST_EMAIL for review")
    parser.add_argument("--send",          action="store_true", help="Send drafts.json to real recipients")
    parser.add_argument("--check-replies", action="store_true", help="Scan inbox for replies and update Excel")

    args = parser.parse_args()

    # Enforce exactly one mode
    modes = [args.draft, args.test, args.send, args.check_replies]
    if sum(modes) == 0:
        parser.print_help()
        print("\n⚠ Specify a mode: --draft, --test, --send, or --check-replies")
        sys.exit(1)
    if sum(modes) > 1:
        print("⚠ Only one mode at a time please.")
        sys.exit(1)

    log.info("=" * 60)
    log.info("AI Email Sender v4")
    log.info("=" * 60)

    if args.draft:
        run_draft()
    elif args.test:
        run_send(test_mode=True)
    elif args.send:
        run_send(test_mode=False)
    elif args.check_replies:
        run_check_replies()


if __name__ == "__main__":
    main()
